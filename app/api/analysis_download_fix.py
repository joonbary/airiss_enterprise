# 🔥 추가: 결과 다운로드 엔드포인트 (수정된 버전)
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@router.get("/download/{job_id}/{format}")
async def download_results(job_id: str, format: str = "excel"):
    """분석 결과 다운로드 (Excel, CSV, JSON)"""
    try:
        logger.info(f"📥 다운로드 요청: {job_id} - 형식: {format}")
        
        db_service = get_db_service()
        if not db_service:
            logger.error("❌ DB 서비스를 사용할 수 없습니다")
            raise HTTPException(status_code=503, detail="데이터베이스 서비스를 사용할 수 없습니다")
        
        await db_service.init_database()
        
        # 작업 및 결과 조회
        job_data = await db_service.get_analysis_job(job_id)
        if not job_data:
            raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
        
        results = await db_service.get_analysis_results(job_id)
        if not results:
            raise HTTPException(status_code=404, detail="분석 결과가 없습니다")
        
        # 결과 데이터 추출 - JSON 파싱 처리 추가
        logger.info(f"📋 결과 데이터 추출 중 - {len(results)}개 레코드")
        result_list = []
        for result in results:
            try:
                # result_data가 문자열인 경우 JSON 파싱
                if isinstance(result.get("result_data"), str):
                    import json
                    result_data = json.loads(result["result_data"])
                else:
                    result_data = result.get("result_data", {})
                result_list.append(result_data)
            except Exception as e:
                logger.error(f"⚠️ 결과 데이터 파싱 오류: {e}")
                continue
        
        if not result_list:
            raise HTTPException(status_code=500, detail="분석 결과 데이터를 파싱할 수 없습니다")
        
        df = pd.DataFrame(result_list)
        logger.info(f"✅ DataFrame 생성 완료: {df.shape}")
        logger.info(f"📊 컬럼 목록: {list(df.columns)}")
        
        # 파일 이름 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"AIRISS_분석결과_{job_id[:8]}_{timestamp}"
        
        if format.lower() == "csv":
            # CSV 다운로드
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"}
            )
            
        elif format.lower() == "json":
            # JSON 다운로드
            json_data = df.to_json(orient='records', force_ascii=False, indent=2)
            
            return StreamingResponse(
                io.BytesIO(json_data.encode('utf-8')),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.json"}
            )
            
        else:  # Excel (기본값)
            # Excel 다운로드 (스타일 적용)
            output = io.BytesIO()
            
            try:
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # 점수 컬럼이 있는지 확인
                    score_column = None
                    for col in ['AIRISS_v4_종합점수', '종합점수', 'overall_score', 'score']:
                        if col in df.columns:
                            score_column = col
                            break
                    
                    # 요약 시트
                    summary_data = {
                        '항목': ['분석일시', '총 분석건수', '평균 점수', '최고 점수', '최저 점수', '분석 모드'],
                        '값': [
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            len(result_list),
                            round(df[score_column].mean(), 1) if score_column and score_column in df.columns else 'N/A',
                            df[score_column].max() if score_column and score_column in df.columns else 'N/A',
                            df[score_column].min() if score_column and score_column in df.columns else 'N/A',
                            job_data.get('analysis_mode', 'hybrid')
                        ]
                    }
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='요약', index=False)
                    
                    # 상세 결과 시트
                    df.to_excel(writer, sheet_name='상세결과', index=False)
                    
                    # 스타일 적용 - 오류 방지를 위해 try-except 추가
                    try:
                        workbook = writer.book
                        
                        # 요약 시트 스타일
                        if '요약' in workbook.sheetnames:
                            summary_sheet = workbook['요약']
                            for row in summary_sheet.iter_rows(min_row=1, max_row=1):
                                for cell in row:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # 상세결과 시트 스타일
                        if '상세결과' in workbook.sheetnames:
                            detail_sheet = workbook['상세결과']
                            # 헤더 스타일
                            for row in detail_sheet.iter_rows(min_row=1, max_row=1):
                                for cell in row:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            # 열 너비 자동 조정 - 오류 방지
                            for column_cells in detail_sheet.columns:
                                try:
                                    # 빈 컬럼 확인
                                    if not column_cells:
                                        continue
                                    
                                    max_length = 0
                                    column_letter = column_cells[0].column_letter if column_cells else 'A'
                                    
                                    for cell in column_cells:
                                        try:
                                            cell_value = str(cell.value) if cell.value is not None else ''
                                            if len(cell_value) > max_length:
                                                max_length = len(cell_value)
                                        except:
                                            continue
                                    
                                    adjusted_width = min(max(max_length + 2, 10), 50)
                                    detail_sheet.column_dimensions[column_letter].width = adjusted_width
                                except Exception as col_error:
                                    logger.warning(f"⚠️ 열 너비 조정 건너뜀: {col_error}")
                                    continue
                    
                    except Exception as style_error:
                        logger.warning(f"⚠️ 스타일 적용 실패 (데이터는 정상): {style_error}")
                
                output.seek(0)
                
                logger.info(f"✅ Excel 파일 생성 완료: {filename_base}.xlsx")
                
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"}
                )
            
            except Exception as excel_error:
                logger.error(f"❌ Excel 생성 오류: {excel_error}")
                logger.error(f"오류 상세: {traceback.format_exc()}")
                
                # Excel 생성 실패 시 CSV로 대체
                logger.info("📋 Excel 생성 실패, CSV로 대체 제공")
                output = io.StringIO()
                df.to_csv(output, index=False, encoding='utf-8-sig')
                output.seek(0)
                
                return StreamingResponse(
                    io.BytesIO(output.getvalue().encode('utf-8-sig')),
                    media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"}
                )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ 다운로드 오류: {e}")
        logger.error(f"오류 상세: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"다운로드 실패: {str(e)}")
