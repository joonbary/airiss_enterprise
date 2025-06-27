# app/api/analysis.py
# AIRISS v4.0 Analysis API - ë¬´í•œ ë¡œë”© í•´ê²° ì™„ë£Œ ë²„ì „
# ğŸ”¥ í•µì‹¬ ìˆ˜ì •: ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì•ˆì •í™” + ì˜ˆì™¸ ì²˜ë¦¬ ê°•í™”

from fastapi import APIRouter, HTTPException, BackgroundTasks, Depends
from pydantic import BaseModel
from typing import Optional, Dict, Any, List
import uuid
import asyncio
import logging
import traceback
from datetime import datetime
import pandas as pd
import numpy as np
import json

# ë¡œê¹… ì„¤ì •
logger = logging.getLogger(__name__)

# ë¼ìš°í„° ìƒì„±
router = APIRouter(prefix="/analysis", tags=["analysis"])

# ğŸ”¥ ì „ì—­ ì¸ìŠ¤í„´ìŠ¤ ì €ì¥ìš©
_db_service = None
_ws_manager = None

def get_db_service():
    """DB ì„œë¹„ìŠ¤ ê°€ì ¸ì˜¤ê¸°"""
    if _db_service is None:
        from app.db.sqlite_service import SQLiteService
        return SQLiteService()
    return _db_service

def get_ws_manager():
    """WebSocket ë§¤ë‹ˆì € ê°€ì ¸ì˜¤ê¸°"""
    if _ws_manager is None:
        from app.core.websocket_manager import ConnectionManager
        return ConnectionManager()
    return _ws_manager

# ğŸ”¥ ì´ˆê¸°í™” í•¨ìˆ˜ (main.pyì—ì„œ í˜¸ì¶œ)
def init_services(db_service, ws_manager):
    """ì„œë¹„ìŠ¤ ì¸ìŠ¤í„´ìŠ¤ ì´ˆê¸°í™”"""
    global _db_service, _ws_manager
    _db_service = db_service
    _ws_manager = ws_manager
    logger.info("âœ… Analysis ëª¨ë“ˆ ì„œë¹„ìŠ¤ ì´ˆê¸°í™” ì™„ë£Œ")

# ì„œë¹„ìŠ¤ì—ì„œ í•˜ì´ë¸Œë¦¬ë“œ ë¶„ì„ê¸° import
try:
    from app.services import HybridAnalyzer
    hybrid_analyzer = HybridAnalyzer()
    logger.info("âœ… ì„œë¹„ìŠ¤ HybridAnalyzer ë¡œë“œ ì„±ê³µ")
except ImportError:
    logger.warning("âš ï¸ ì„œë¹„ìŠ¤ HybridAnalyzer ë¡œë“œ ì‹¤íŒ¨, ë¡œì»¬ ì •ì˜ ì‚¬ìš©")
    hybrid_analyzer = None

# ğŸ†• v3.0 AIRISS 8ëŒ€ ì˜ì—­ ì™„ì „ ë³´ì¡´
# ğŸ”¥ ê¸°ì¡´ AIRISS_FRAMEWORK ë¶€ë¶„ì„ ì°¾ì•„ì„œ ì•„ë˜ ì½”ë“œë¡œ ì™„ì „ êµì²´í•˜ì„¸ìš”
# íŒŒì¼: app/api/analysis.py (ì•½ 25~140ë²ˆì§¸ ì¤„)

# ğŸ†• ì°©ìˆ˜ë³´ê³ ì„œ ì™„ì „ ë°˜ì˜ - AIRISS 8ëŒ€ ì˜ì—­ 
AIRISS_FRAMEWORK = {
    "ì—…ë¬´ì„±ê³¼": {
        "keywords": {
            "positive": [
                "ìš°ìˆ˜", "íƒì›”", "ë›°ì–´ë‚¨", "ì„±ê³¼", "ë‹¬ì„±", "ì™„ë£Œ", "ì„±ê³µ", "íš¨ìœ¨", "ìƒì‚°ì ", 
                "ëª©í‘œë‹¬ì„±", "ì´ˆê³¼ë‹¬ì„±", "í’ˆì§ˆ", "ì •í™•", "ì‹ ì†", "ì™„ë²½", "ì „ë¬¸ì ", "ì²´ê³„ì ",
                "ì„±ê³¼ê°€", "ê²°ê³¼ë¥¼", "ì‹¤ì ì´", "ì™„ì„±ë„", "ë§Œì¡±ë„", "ì‚°ì¶œë¬¼", "ì•„ì›ƒí’‹",
                "ì—…ë¬´ì™„ë£Œ", "í”„ë¡œì íŠ¸", "ë³´ê³ ì„œ", "ê²°ê³¼ë¬¼", "deliverable"
            ],
            "negative": [
                "ë¶€ì¡±", "ë¯¸í¡", "ì§€ì—°", "ì‹¤íŒ¨", "ë¬¸ì œ", "ì˜¤ë¥˜", "ëŠ¦ìŒ", "ë¹„íš¨ìœ¨", 
                "ëª©í‘œë¯¸ë‹¬", "í’ˆì§ˆì €í•˜", "ë¶€ì •í™•", "ë¯¸ì™„ì„±", "ë¶€ì‹¤", "ê°œì„ ", "ë³´ì™„",
                "ì§€ì²´", "ì‚°ì¶œë¬¼ë¶€ì¡±", "ê²°ê³¼ë¬¼ë¬¸ì œ"
            ]
        },
        "weight": 0.20,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: ì—…ë¬´ ì‚°ì¶œë¬¼ 20%
        "description": "ì—…ë¬´ ì‚°ì¶œë¬¼ì˜ ì–‘ê³¼ ì§ˆì  ìˆ˜ì¤€",
        "color": "#FF5722",
        "icon": "ğŸ“Š"
    },
    "KPIë‹¬ì„±": {
        "keywords": {
            "positive": [
                "KPIë‹¬ì„±", "ì§€í‘œë‹¬ì„±", "ëª©í‘œì´ˆê³¼", "ì„±ê³¼ìš°ìˆ˜", "ì‹¤ì ìš°ìˆ˜", "ë§¤ì¶œì¦ê°€", 
                "íš¨ìœ¨í–¥ìƒ", "ìƒì‚°ì„±í–¥ìƒ", "ìˆ˜ì¹˜ë‹¬ì„±", "ì„±ì¥", "ê°œì„ ", "ë‹¬ì„±ë¥ ", "ì´ˆê³¼",
                "ROA", "ROE", "ìˆ˜ìµë¥ ", "ì—¬ì‹ ", "ê³ ê°ë§Œì¡±ë„", "ì—°ì²´ìœ¨ê°œì„ ", "ë¦¬ìŠ¤í¬ê´€ë¦¬",
                "í•µì‹¬ì§€í‘œ", "ì •ëŸ‰ëª©í‘œ", "ìˆ˜ì¹˜ëª©í‘œ", "ì‹¤ì ì´ˆê³¼", "ì§€í‘œê°œì„ "
            ],
            "negative": [
                "KPIë¯¸ë‹¬", "ëª©í‘œë¯¸ë‹¬", "ì‹¤ì ë¶€ì§„", "ë§¤ì¶œê°ì†Œ", "íš¨ìœ¨ì €í•˜", 
                "ìƒì‚°ì„±ì €í•˜", "ìˆ˜ì¹˜ë¶€ì¡±", "í•˜ë½", "í‡´ë³´", "ë¯¸ë‹¬", "ì†ì‹¤",
                "ì§€í‘œì•…í™”", "ëª©í‘œë¶€ì¡±", "ì‹¤ì ì €ì¡°"
            ]
        },
        "weight": 0.30,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: KPI 30%
        "description": "í•µì‹¬ì„±ê³¼ì§€í‘œ ë‹¬ì„±ë„ ë° ì •ëŸ‰ì  ê¸°ì—¬",
        "color": "#4A4A4A",
        "icon": "ğŸ¯"
    },
    "íƒœë„ë§ˆì¸ë“œì…‹": {
        "keywords": {
            "positive": [
                "ì ê·¹ì ", "ê¸ì •ì ", "ì—´ì •", "ì„±ì‹¤", "ì±…ì„ê°", "ì§„ì·¨ì ", "í˜‘ì¡°ì ", 
                "ì„±ì¥ì§€í–¥", "í•™ìŠµì˜ì§€", "ë„ì „ì •ì‹ ", "ì£¼ì¸ì˜ì‹", "í—Œì‹ ", "ì—´ì‹¬íˆ", "ë…¸ë ¥",
                "ë³€í™”ìˆ˜ìš©", "íšŒë³µíƒ„ë ¥ì„±", "ê·¼ë¬´íƒœë„", "ì§„ì •ì„±", "ìì„¸", "ë§ˆì¸ë“œì…‹",
                "ì˜ìš•", "ë™ê¸°", "ëª°ì…", "ì§‘ì¤‘", "ì„±ì‹¤ì„±", "ê·¼ë©´", "ë¶€ì§€ëŸ°í•¨"
            ],
            "negative": [
                "ì†Œê·¹ì ", "ë¶€ì •ì ", "ë¬´ê´€ì‹¬", "ë¶ˆì„±ì‹¤", "íšŒí”¼", "ëƒ‰ì†Œì ", 
                "ë¹„í˜‘ì¡°ì ", "ì•ˆì£¼", "í˜„ìƒìœ ì§€", "ìˆ˜ë™ì ", "íƒœë„ë¬¸ì œ", "ë§ˆì¸ë“œë¶€ì¡±",
                "ë³€í™”ê±°ë¶€", "ì˜ìš•ì—†ìŒ", "ë¬´ê¸°ë ¥", "ëƒ‰ë‹´", "ë¶ˆë§Œ"
            ]
        },
        "weight": 0.10,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: íƒœë„ ë° ë§ˆì¸ë“œì…‹ 10%
        "description": "ì¼ì— ëŒ€í•œ íƒœë„ì™€ ë³€í™” ìˆ˜ìš© ë§ˆì¸ë“œì…‹",
        "color": "#F89C26",
        "icon": "ğŸ§ "
    },
    "ì»¤ë®¤ë‹ˆì¼€ì´ì…˜ì—­ëŸ‰": {
        "keywords": {
            "positive": [
                "ëª…í™•", "ì •í™•", "ì‹ ì†", "ì¹œì ˆ", "ê²½ì²­", "ì†Œí†µ", "ì „ë‹¬", "ì´í•´", 
                "ì„¤ë“", "í˜‘ì˜", "ì¡°ìœ¨", "ê³µìœ ", "íˆ¬ëª…", "ê°œë°©ì ", "ì˜ì‚¬ì†Œí†µ", "ì›í™œ",
                "ì‘ë‹µì†ë„", "ëª…í™•ì„±", "í†¤", "ì˜í–¥ë ¥", "ê³ ê°ì†Œí†µ", "ë‚´ë¶€ì†Œí†µ",
                "ëŒ€í™”", "ì§ˆë¬¸", "ë‹µë³€", "í”¼ë“œë°±", "ë³´ê³ ", "ë°œí‘œ", "ì„¤ëª…"
            ],
            "negative": [
                "ë¶ˆëª…í™•", "ì§€ì—°", "ë¬´ì‹œ", "ì˜¤í•´", "ë‹¨ì ˆ", "ì¹¨ë¬µ", "íšŒí”¼", 
                "ë…ë‹¨", "ì¼ë°©ì ", "íì‡„ì ", "ì†Œí†µë¶€ì¡±", "ì „ë‹¬ë ¥ë¶€ì¡±", "ì‘ë‹µì§€ì—°",
                "ì†Œí†µë¬¸ì œ", "ì˜ì‚¬ì „ë‹¬", "ì»¤ë®¤ë‹ˆì¼€ì´ì…˜ë¶€ì¡±"
            ]
        },
        "weight": 0.10,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: ì»¤ë®¤ë‹ˆì¼€ì´ì…˜ ì—­ëŸ‰ 10%
        "description": "ì˜ì‚¬ì†Œí†µ íš¨ê³¼ì„±ê³¼ ê´€ê³„í˜•ì„± ëŠ¥ë ¥",
        "color": "#B3B3B3",
        "icon": "ğŸ’¬"
    },
    "ë¦¬ë”ì‹­í˜‘ì—…ì—­ëŸ‰": {
        "keywords": {
            "positive": [
                "ë¦¬ë”ì‹­", "íŒ€ì›Œí¬", "í˜‘ì—…", "ì§€ì›", "ë©˜í† ë§", "ë™ê¸°ë¶€ì—¬", "ì¡°ìœ¨", 
                "í™”í•©", "íŒ€ë¹Œë”©", "ìœ„ì„", "ì½”ì¹­", "ì˜í–¥ë ¥", "í˜‘ë ¥", "íŒ€í”Œë ˆì´",
                "íŒ€ì„±ê³¼", "ë¶€í•˜ì§ì›", "ë™ë£Œì§€ì›", "ê°ˆë“±í•´ê²°", "í•©ì˜ë„ì¶œ",
                "ê³µë™ì‘ì—…", "í˜‘ì¡°", "ì¡°í™”", "ì‹œë„ˆì§€", "ìƒí˜¸ë³´ì™„"
            ],
            "negative": [
                "ë…ë‹¨", "ê°ˆë“±", "ë¹„í˜‘ì¡°", "ì†Œì™¸", "ë¶„ì—´", "ëŒ€ë¦½", "ì´ê¸°ì£¼ì˜", 
                "ë°©í•´", "ë¬´ê´€ì‹¬", "ê³ ë¦½", "ê°œì¸ì£¼ì˜", "íŒ€ì›Œí¬ë¶€ì¡±",
                "í˜‘ì—…ë¬¸ì œ", "ë¦¬ë”ì‹­ë¶€ì¡±", "íŒ€í™”í•©ì €í•´"
            ]
        },
        "weight": 0.10,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: ë¦¬ë”ì‹­ & í˜‘ì—… ì—­ëŸ‰ 10%
        "description": "ë¦¬ë”ì‹­ ë°œíœ˜ì™€ í˜‘ì—… ì´‰ì§„ ëŠ¥ë ¥",
        "color": "#FF8A50",
        "icon": "ğŸ‘¥"
    },
    "ì§€ì‹ì „ë¬¸ì„±": {
        "keywords": {
            "positive": [
                "ì „ë¬¸", "ìˆ™ë ¨", "ê¸°ìˆ ", "ì§€ì‹", "í•™ìŠµ", "ë°œì „", "ì—­ëŸ‰", "ëŠ¥ë ¥", 
                "ì„±ì¥", "í–¥ìƒ", "ìŠµë“", "ê°œë°œ", "ì „ë¬¸ì„±", "ë…¸í•˜ìš°", "ìŠ¤í‚¬", "ê²½í—˜",
                "ìê²©ì¦", "êµìœ¡", "ì—°ìˆ˜", "AIì—­ëŸ‰", "ë””ì§€í„¸ì—­ëŸ‰", "ê¸ˆìœµì „ë¬¸ì„±",
                "ì „ë¬¸ì§€ì‹", "ê¸°ìˆ ë ¥", "ì‹¤ë ¥", "ìˆ™ë ¨ë„", "ì „ë¬¸ë¶„ì•¼", "ê¹Šì´"
            ],
            "negative": [
                "ë¯¸ìˆ™", "ë¶€ì¡±", "ë‚™í›„", "ë¬´ì§€", "ì •ì²´", "í‡´ë³´", "ë¬´ëŠ¥ë ¥", 
                "ê¸°ì´ˆë¶€ì¡±", "ì—­ëŸ‰ë¶€ì¡±", "ì‹¤ë ¥ë¶€ì¡±", "í•™ìŠµê±°ë¶€", "ì§€ì‹ë¶€ì¡±",
                "ì „ë¬¸ì„±ë¶€ì¡±", "ê¸°ìˆ ë¶€ì¡±", "ê²½í—˜ë¶€ì¡±"
            ]
        },
        "weight": 0.10,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: ì§€ì‹ & ì „ë¬¸ì„± 10%
        "description": "ì§ë¬´ ì „ë¬¸ì„±ê³¼ ì§€ì† í•™ìŠµ ëŠ¥ë ¥",
        "color": "#6A6A6A",
        "icon": "ğŸ“š"
    },
    "ë¼ì´í”„ìŠ¤íƒ€ì¼ê±´ê°•": {  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ ì™„ì „ ë°˜ì˜ (ê¸°ì¡´ "ì°½ì˜í˜ì‹ " ëŒ€ì²´)
        "keywords": {
            "positive": [
                "ê±´ê°•", "í™œë ¥", "ì—ë„ˆì§€", "ì›Œë¼ë°¸", "ê· í˜•", "ì›°ë¹™", "ìš´ë™", "ëª…ìƒ",
                "ìŠ¤íŠ¸ë ˆìŠ¤ê´€ë¦¬", "ìˆ˜ë©´", "í™œê¸°", "ì»¨ë””ì…˜", "ì²´ë ¥", "ëª°ì…",
                "ì›°ë¹™í”„ë¡œê·¸ë¨", "ê±´ê°•ê´€ë¦¬", "ìƒí™œìŠµê´€", "ì •ì‹ ê±´ê°•", "ì²´ë ¥ê´€ë¦¬",
                "ë°¸ëŸ°ìŠ¤", "íœ´ì‹", "ì¬ì¶©ì „", "í™œë™ì ", "ê±´ê°•ìƒíƒœ"
            ],
            "negative": [
                "í”¼ë¡œ", "ìŠ¤íŠ¸ë ˆìŠ¤", "ë²ˆì•„ì›ƒ", "ê³¼ë¡œ", "ë¶ˆê· í˜•", "ê±´ê°•ì•…í™”",
                "ë³‘ê°€", "ê²°ê·¼", "ì»¨ë””ì…˜ë‚œì¡°", "ì§‘ì¤‘ë ¥ì €í•˜", "ë¬´ê¸°ë ¥", "ì†Œì§„",
                "ë¶ˆê±´ê°•", "ì²´ë ¥ì €í•˜", "ìŠ¤íŠ¸ë ˆìŠ¤ê³¼ë‹¤", "ê³¼ë¡œëˆ„ì "
            ]
        },
        "weight": 0.05,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: ë¼ì´í”„ìŠ¤íƒ€ì¼ & ê±´ê°• 5%
        "description": "ì—…ë¬´ ì§€ì†ì„±ê³¼ ëª°ì…ë„ì— ì˜í–¥í•˜ëŠ” ê±´ê°•ê³¼ ì›°ë¹™",
        "color": "#4CAF50",
        "icon": "ğŸ’ª"
    },
    "ìœ¤ë¦¬ì‚¬ì™¸í–‰ë™": {  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ ì™„ì „ ë°˜ì˜ (ê¸°ì¡´ "ì¡°ì§ì ì‘" ëŒ€ì²´)
        "keywords": {
            "positive": [
                "ìœ¤ë¦¬", "ì‹ ë¢°", "ì„±ì‹¤", "ì •ì§", "íˆ¬ëª…", "ì¤€ë²•", "ê·œì •ì¤€ìˆ˜", "ì²­ë ´",
                "ë´‰ì‚¬", "ì‚¬íšŒê³µí—Œ", "ì§€ì—­ì‚¬íšŒ", "CSR", "ëª¨ë²”", "í’ˆìœ„", "í’ˆê²©",
                "ì„ì§ì›ìœ¤ë¦¬ê°•ë ¹", "ì»´í”Œë¼ì´ì–¸ìŠ¤", "ë¦¬ìŠ¤í¬ê´€ë¦¬", "í‰íŒ", "ì‹ ë¢°ì„±",
                "ë„ë•ì ", "ì–‘ì‹¬ì ", "ì±…ì„ê°", "ì‚¬íšŒì ì±…ì„"
            ],
            "negative": [
                "ìœ„ë°˜", "ë¹„ìœ¤ë¦¬", "ë¶ˆë²•", "ë¶€ì •", "ìŠ¤ìº”ë“¤", "ë…¼ë€", "ë¬¸ì œí–‰ë™",
                "ê·œì •ìœ„ë°˜", "ë¦¬ìŠ¤í¬", "í‰íŒì†ìƒ", "ì‹ ë¢°ì‹¤ì¶”", "ìœ„ë²•í–‰ìœ„",
                "SNSë…¼ë€", "í˜ì˜¤ë°œì–¸", "ë¶€ì ì ˆí–‰ë™", "ìœ¤ë¦¬ë¬¸ì œ", "ë„ë•ì í•´ì´",
                "ì‚¬íšŒì ë¬¼ì˜", "ë¹„ë¦¬", "ë¶€íŒ¨"
            ]
        },
        "weight": 0.05,  # ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ: ì‚¬ì™¸ í–‰ë™ ë° ìœ¤ë¦¬ 5%
        "description": "ì¡°ì§ ì‹ ë¢°ë„ì™€ í‰íŒì— ì˜í–¥í•˜ëŠ” ìœ¤ë¦¬ì„±ê³¼ ì‚¬ì™¸í–‰ë™",
        "color": "#9E9E9E",
        "icon": "âš–ï¸"
    }
}

# ğŸ”¥ ì°©ìˆ˜ë³´ê³ ì„œ ê°€ì¤‘ì¹˜ ê²€ì¦
OFFICIAL_WEIGHTS = {
    "ì—…ë¬´ì„±ê³¼": 0.20,
    "KPIë‹¬ì„±": 0.30,
    "íƒœë„ë§ˆì¸ë“œì…‹": 0.10,
    "ì»¤ë®¤ë‹ˆì¼€ì´ì…˜ì—­ëŸ‰": 0.10,
    "ë¦¬ë”ì‹­í˜‘ì—…ì—­ëŸ‰": 0.10,
    "ì§€ì‹ì „ë¬¸ì„±": 0.10,
    "ë¼ì´í”„ìŠ¤íƒ€ì¼ê±´ê°•": 0.05,
    "ìœ¤ë¦¬ì‚¬ì™¸í–‰ë™": 0.05
}

# ê°€ì¤‘ì¹˜ í•©ê³„ ê²€ì¦ (ë°˜ë“œì‹œ 1.0ì´ì–´ì•¼ í•¨)
total_weight = sum(OFFICIAL_WEIGHTS.values())
assert total_weight == 1.0, f"âŒ ê°€ì¤‘ì¹˜ í•©ê³„ ì˜¤ë¥˜: {total_weight}"
logger.info(f"âœ… ì°©ìˆ˜ë³´ê³ ì„œ ê¸°ì¤€ ê°€ì¤‘ì¹˜ ê²€ì¦ ì™„ë£Œ: ì´í•© {total_weight}")

# ğŸ†• ì •ëŸ‰ë°ì´í„° ë¶„ì„ê¸° (v3.0 ì™„ì „ ë³´ì¡´)
class QuantitativeAnalyzer:
    """í‰ê°€ë“±ê¸‰, ì ìˆ˜ ë“± ì •ëŸ‰ë°ì´í„° ë¶„ì„ ì „ìš© í´ë˜ìŠ¤"""
    
    def __init__(self):
        self.grade_mappings = self.setup_grade_mappings()
        self.score_weights = self.setup_score_weights()
        logger.info("âœ… ì •ëŸ‰ë°ì´í„° ë¶„ì„ê¸° ì´ˆê¸°í™” ì™„ë£Œ")
    
    def setup_grade_mappings(self) -> Dict[str, Dict]:
        """ë‹¤ì–‘í•œ í‰ê°€ë“±ê¸‰ í˜•ì‹ì„ ì ìˆ˜ë¡œ ë³€í™˜í•˜ëŠ” ë§¤í•‘ í…Œì´ë¸”"""
        return {
            # 5ë‹¨ê³„ ë“±ê¸‰
            'S': 100, 'A': 85, 'B': 70, 'C': 55, 'D': 40,
            
            # ì˜ë¬¸ ë“±ê¸‰  
            'A+': 100, 'A': 95, 'A-': 90,
            'B+': 85, 'B': 80, 'B-': 75,
            'C+': 70, 'C': 65, 'C-': 60,
            'D+': 55, 'D': 50, 'D-': 45,
            'F': 30,
            
            # ìˆ«ì ë“±ê¸‰
            '1': 100, '2': 80, '3': 60, '4': 40, '5': 20,
            '1ê¸‰': 100, '2ê¸‰': 80, '3ê¸‰': 60, '4ê¸‰': 40, '5ê¸‰': 20,
            
            # í•œê¸€ ë“±ê¸‰
            'ìš°ìˆ˜': 90, 'ì–‘í˜¸': 75, 'ë³´í†µ': 60, 'ë¯¸í¡': 45, 'ë¶€ì¡±': 30,
            'ìµœìš°ìˆ˜': 100, 'ìƒ': 85, 'ì¤‘': 65, 'í•˜': 45,
            
            # ë°±ë¶„ìœ„/í¼ì„¼íŠ¸
            'ìƒìœ„10%': 95, 'ìƒìœ„20%': 85, 'ìƒìœ„30%': 75, 
            'ìƒìœ„50%': 65, 'í•˜ìœ„50%': 50, 'í•˜ìœ„30%': 35,
            
            # OKê¸ˆìœµê·¸ë£¹ ë§ì¶¤ ë“±ê¸‰
            'OKâ˜…â˜…â˜…': 100, 'OKâ˜…â˜…': 90, 'OKâ˜…': 80, 
            'OK A': 75, 'OK B+': 70, 'OK B': 65, 'OK C': 55, 'OK D': 40
        }
    
    def setup_score_weights(self) -> Dict[str, float]:
        """ì •ëŸ‰ ë°ì´í„° í•­ëª©ë³„ ê°€ì¤‘ì¹˜ ì„¤ì •"""
        return {
            'performance_grade': 0.30,
            'kpi_score': 0.25,
            'competency_grade': 0.20,
            'attendance_score': 0.10,
            'training_score': 0.10,
            'certificate_score': 0.05
        }
    
    def extract_quantitative_data(self, row: pd.Series) -> Dict[str, Any]:
        """í–‰ ë°ì´í„°ì—ì„œ ì •ëŸ‰ì  ìš”ì†Œ ì¶”ì¶œ"""
        quant_data = {}
        
        for col_name, value in row.items():
            col_lower = str(col_name).lower()
            
            if any(keyword in col_lower for keyword in ['ì ìˆ˜', 'score', 'í‰ì ', 'rating']):
                quant_data[f'score_{col_name}'] = self.normalize_score(value)
            elif any(keyword in col_lower for keyword in ['ë“±ê¸‰', 'grade', 'í‰ê°€', 'level']):
                quant_data[f'grade_{col_name}'] = self.convert_grade_to_score(value)
            elif any(keyword in col_lower for keyword in ['ë‹¬ì„±ë¥ ', 'ë¹„ìœ¨', 'rate', '%', 'percent']):
                quant_data[f'rate_{col_name}'] = self.normalize_percentage(value)
            elif any(keyword in col_lower for keyword in ['íšŸìˆ˜', 'ê±´ìˆ˜', 'count', 'íšŒ', 'ë²ˆ']):
                quant_data[f'count_{col_name}'] = self.normalize_count(value)
                
        return quant_data
    
    def convert_grade_to_score(self, grade_value) -> float:
        """ë“±ê¸‰ì„ ì ìˆ˜ë¡œ ë³€í™˜"""
        if pd.isna(grade_value) or grade_value == '':
            return 50.0
        
        grade_str = str(grade_value).strip().upper()
        
        if grade_str in self.grade_mappings:
            return float(self.grade_mappings[grade_str])
        
        try:
            score = float(grade_str)
            if 0 <= score <= 100:
                return score
            elif 0 <= score <= 5:
                return (score - 1) * 25
            elif 0 <= score <= 10:
                return score * 10
        except ValueError:
            pass
        
        if 'ìš°ìˆ˜' in grade_str or 'excellent' in grade_str.lower():
            return 90.0
        elif 'ì–‘í˜¸' in grade_str or 'good' in grade_str.lower():
            return 75.0
        elif 'ë³´í†µ' in grade_str or 'average' in grade_str.lower():
            return 60.0
        elif 'ë¯¸í¡' in grade_str or 'poor' in grade_str.lower():
            return 45.0
        
        return 50.0
    
    def normalize_score(self, score_value) -> float:
        """ì ìˆ˜ ê°’ ì •ê·œí™” (0-100 ë²”ìœ„ë¡œ)"""
        if pd.isna(score_value) or score_value == '':
            return 50.0
        
        try:
            score = float(str(score_value).replace('%', '').replace('ì ', ''))
            
            if 0 <= score <= 1:
                return score * 100
            elif 0 <= score <= 5:
                return (score - 1) * 25
            elif 0 <= score <= 10:
                return score * 10
            elif 0 <= score <= 100:
                return score
            else:
                return max(0, min(100, score))
                
        except (ValueError, TypeError):
            return 50.0
    
    def normalize_percentage(self, percent_value) -> float:
        """ë°±ë¶„ìœ¨ ì •ê·œí™”"""
        if pd.isna(percent_value) or percent_value == '':
            return 50.0
        
        try:
            percent_str = str(percent_value).replace('%', '').replace('í¼ì„¼íŠ¸', '')
            percent = float(percent_str)
            
            if 0 <= percent <= 1:
                return percent * 100
            elif 0 <= percent <= 100:
                return percent
            else:
                return max(0, min(100, percent))
                
        except (ValueError, TypeError):
            return 50.0
    
    def normalize_count(self, count_value) -> float:
        """íšŸìˆ˜/ê±´ìˆ˜ë¥¼ ì ìˆ˜ë¡œ ë³€í™˜"""
        if pd.isna(count_value) or count_value == '':
            return 50.0
        
        try:
            count = float(str(count_value).replace('íšŒ', '').replace('ê±´', '').replace('ë²ˆ', ''))
            
            if count <= 0:
                return 30.0
            elif count <= 2:
                return 50.0
            elif count <= 5:
                return 70.0
            elif count <= 10:
                return 85.0
            else:
                return 95.0
                
        except (ValueError, TypeError):
            return 50.0
    
    def calculate_quantitative_score(self, quant_data: Dict[str, float]) -> Dict[str, Any]:
        """ì •ëŸ‰ ë°ì´í„°ë“¤ì„ ì¢…í•©í•˜ì—¬ ìµœì¢… ì ìˆ˜ ê³„ì‚°"""
        if not quant_data:
            return {
                "quantitative_score": 50.0,
                "confidence": 0.0,
                "contributing_factors": {},
                "data_quality": "ì—†ìŒ"
            }
        
        total_score = 0.0
        total_weight = 0.0
        contributing_factors = {}
        
        for data_key, score in quant_data.items():
            if 'grade_' in data_key:
                weight = 0.4
            elif 'score_' in data_key:
                weight = 0.3
            elif 'rate_' in data_key:
                weight = 0.2
            else:
                weight = 0.1
            
            total_score += score * weight
            total_weight += weight
            contributing_factors[data_key] = {
                "score": round(score, 1),
                "weight": weight,
                "contribution": round(score * weight, 1)
            }
        
        if total_weight > 0:
            final_score = total_score / total_weight
            confidence = min(total_weight * 20, 100)
        else:
            final_score = 50.0
            confidence = 0.0
        
        data_count = len(quant_data)
        if data_count >= 5:
            data_quality = "ë†’ìŒ"
        elif data_count >= 3:
            data_quality = "ì¤‘ê°„"
        elif data_count >= 1:
            data_quality = "ë‚®ìŒ"
        else:
            data_quality = "ì—†ìŒ"
        
        return {
            "quantitative_score": round(final_score, 1),
            "confidence": round(confidence, 1),
            "contributing_factors": contributing_factors,
            "data_quality": data_quality,
            "data_count": data_count
        }

# ğŸ†• í…ìŠ¤íŠ¸ ë¶„ì„ê¸° (v3.0 ì™„ì „ ë³´ì¡´)
class AIRISSTextAnalyzer:
    def __init__(self):
        self.framework = AIRISS_FRAMEWORK
        self.openai_available = False
        self.openai = None
        try:
            import openai
            self.openai = openai
            self.openai_available = True
            logger.info("âœ… OpenAI ëª¨ë“ˆ ë¡œë“œ ì„±ê³µ")
        except ImportError:
            logger.warning("âš ï¸ OpenAI ëª¨ë“ˆ ì—†ìŒ - í‚¤ì›Œë“œ ë¶„ì„ë§Œ ê°€ëŠ¥")
    
    def analyze_text(self, text: str, dimension: str) -> Dict[str, Any]:
        """í…ìŠ¤íŠ¸ ë¶„ì„í•˜ì—¬ ì ìˆ˜ ì‚°ì¶œ"""
        if not text or text.lower() in ['nan', 'null', '', 'none']:
            return {"score": 50, "confidence": 0, "signals": {"positive": 0, "negative": 0, "positive_words": [], "negative_words": []}}
        
        keywords = self.framework[dimension]["keywords"]
        text_lower = text.lower()
        
        positive_matches = []
        negative_matches = []
        
        for word in keywords["positive"]:
            if word in text_lower:
                positive_matches.append(word)
        
        for word in keywords["negative"]:
            if word in text_lower:
                negative_matches.append(word)
        
        positive_count = len(positive_matches)
        negative_count = len(negative_matches)
        
        base_score = 50
        positive_boost = min(positive_count * 8, 45)
        negative_penalty = min(negative_count * 10, 40)
        
        text_length = len(text)
        if text_length > 50:
            length_bonus = min((text_length - 50) / 100 * 5, 10)
        else:
            length_bonus = 0
        
        final_score = base_score + positive_boost - negative_penalty + length_bonus
        final_score = max(10, min(100, final_score))
        
        total_signals = positive_count + negative_count
        base_confidence = min(total_signals * 12, 80)
        length_confidence = min(text_length / 20, 20)
        confidence = min(base_confidence + length_confidence, 100)
        
        return {
            "score": round(final_score, 1),
            "confidence": round(confidence, 1),
            "signals": {
                "positive": positive_count,
                "negative": negative_count,
                "positive_words": positive_matches[:5],
                "negative_words": negative_matches[:5]
            }
        }
    
    def calculate_overall_score(self, dimension_scores: Dict[str, float]) -> Dict[str, Any]:
        """ì¢…í•© ì ìˆ˜ ê³„ì‚°"""
        weighted_sum = 0
        total_weight = 0
        
        for dimension, score in dimension_scores.items():
            if dimension in self.framework:
                weight = self.framework[dimension]["weight"]
                weighted_sum += score * weight
                total_weight += weight
        
        overall_score = weighted_sum / total_weight if total_weight > 0 else 50
        
        if overall_score >= 95:
            grade = "OKâ˜…â˜…â˜…"
            grade_desc = "ìµœìš°ìˆ˜ ë“±ê¸‰ (TOP 1%)"
            percentile = "ìƒìœ„ 1%"
        elif overall_score >= 90:
            grade = "OKâ˜…â˜…"
            grade_desc = "ìš°ìˆ˜ ë“±ê¸‰ (TOP 5%)"
            percentile = "ìƒìœ„ 5%"
        elif overall_score >= 85:
            grade = "OKâ˜…"
            grade_desc = "ìš°ìˆ˜+ ë“±ê¸‰ (TOP 10%)"
            percentile = "ìƒìœ„ 10%"
        elif overall_score >= 80:
            grade = "OK A"
            grade_desc = "ì–‘í˜¸ ë“±ê¸‰ (TOP 20%)"
            percentile = "ìƒìœ„ 20%"
        elif overall_score >= 75:
            grade = "OK B+"
            grade_desc = "ì–‘í˜¸- ë“±ê¸‰ (TOP 30%)"
            percentile = "ìƒìœ„ 30%"
        elif overall_score >= 70:
            grade = "OK B"
            grade_desc = "ë³´í†µ ë“±ê¸‰ (TOP 40%)"
            percentile = "ìƒìœ„ 40%"
        elif overall_score >= 60:
            grade = "OK C"
            grade_desc = "ê°œì„ í•„ìš” ë“±ê¸‰ (TOP 60%)"
            percentile = "ìƒìœ„ 60%"
        else:
            grade = "OK D"
            grade_desc = "ì§‘ì¤‘ê°œì„  ë“±ê¸‰ (í•˜ìœ„ 40%)"
            percentile = "í•˜ìœ„ 40%"
        
        return {
            "overall_score": round(overall_score, 1),
            "grade": grade,
            "grade_description": grade_desc,
            "percentile": percentile,
            "weighted_scores": dimension_scores
        }

# ğŸ†• í•˜ì´ë¸Œë¦¬ë“œ ë¶„ì„ê¸° (v3.0 ì™„ì „ ë³´ì¡´)
class AIRISSHybridAnalyzer:
    """í…ìŠ¤íŠ¸ ë¶„ì„ + ì •ëŸ‰ ë¶„ì„ í†µí•© í´ë˜ìŠ¤"""
    
    def __init__(self):
        self.text_analyzer = AIRISSTextAnalyzer()
        self.quantitative_analyzer = QuantitativeAnalyzer()
        
        self.hybrid_weights = {
            'text_analysis': 0.6,
            'quantitative_analysis': 0.4
        }
        
        logger.info("âœ… AIRISS v4.0 í•˜ì´ë¸Œë¦¬ë“œ ë¶„ì„ê¸° ì´ˆê¸°í™” ì™„ë£Œ")
    
    def comprehensive_analysis(self, uid: str, opinion: str, row_data: pd.Series) -> Dict[str, Any]:
        """ì¢…í•© ë¶„ì„: í…ìŠ¤íŠ¸ + ì •ëŸ‰ ë°ì´í„°"""
        
        # 1. í…ìŠ¤íŠ¸ ë¶„ì„
        text_results = {}
        for dimension in AIRISS_FRAMEWORK.keys():
            text_results[dimension] = self.text_analyzer.analyze_text(opinion, dimension)
        
        text_overall = self.text_analyzer.calculate_overall_score(
            {dim: result["score"] for dim, result in text_results.items()}
        )
        
        # 2. ì •ëŸ‰ ë°ì´í„° ë¶„ì„
        quant_data = self.quantitative_analyzer.extract_quantitative_data(row_data)
        quant_results = self.quantitative_analyzer.calculate_quantitative_score(quant_data)
        
        # 3. í•˜ì´ë¸Œë¦¬ë“œ ì ìˆ˜ ê³„ì‚°
        text_weight = self.hybrid_weights['text_analysis']
        quant_weight = self.hybrid_weights['quantitative_analysis']
        
        if quant_results["data_quality"] == "ì—†ìŒ":
            text_weight = 0.8
            quant_weight = 0.2
        elif quant_results["data_quality"] == "ë‚®ìŒ":
            text_weight = 0.7
            quant_weight = 0.3
        
        hybrid_score = (text_overall["overall_score"] * text_weight + 
                       quant_results["quantitative_score"] * quant_weight)
        
        # 4. í†µí•© ì‹ ë¢°ë„ ê³„ì‚°
        hybrid_confidence = (text_overall.get("confidence", 70) * text_weight + 
                           quant_results["confidence"] * quant_weight)
        
        # 5. í•˜ì´ë¸Œë¦¬ë“œ ë“±ê¸‰ ì‚°ì •
        hybrid_grade_info = self.calculate_hybrid_grade(hybrid_score)
        
        return {
            "text_analysis": {
                "overall_score": text_overall["overall_score"],
                "grade": text_overall["grade"],
                "dimension_scores": {dim: result["score"] for dim, result in text_results.items()},
                "dimension_details": text_results
            },
            "quantitative_analysis": quant_results,
            "hybrid_analysis": {
                "overall_score": round(hybrid_score, 1),
                "grade": hybrid_grade_info["grade"],
                "grade_description": hybrid_grade_info["grade_description"],
                "percentile": hybrid_grade_info["percentile"],
                "confidence": round(hybrid_confidence, 1),
                "analysis_composition": {
                    "text_weight": round(text_weight * 100, 1),
                    "quantitative_weight": round(quant_weight * 100, 1)
                }
            },
            "analysis_metadata": {
                "uid": uid,
                "analysis_version": "AIRISS v4.0",
                "analysis_timestamp": datetime.now().isoformat(),
                "data_sources": {
                    "text_available": bool(opinion and opinion.strip()),
                    "quantitative_available": bool(quant_data),
                    "quantitative_data_quality": quant_results["data_quality"]
                }
            }
        }
    
    def calculate_hybrid_grade(self, score: float) -> Dict[str, str]:
        """í•˜ì´ë¸Œë¦¬ë“œ ì ìˆ˜ë¥¼ OKë“±ê¸‰ìœ¼ë¡œ ë³€í™˜"""
        if score >= 95:
            return {
                "grade": "OKâ˜…â˜…â˜…",
                "grade_description": "ìµœìš°ìˆ˜ ë“±ê¸‰ (TOP 1%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 1%"
            }
        elif score >= 90:
            return {
                "grade": "OKâ˜…â˜…",
                "grade_description": "ìš°ìˆ˜ ë“±ê¸‰ (TOP 5%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 5%"
            }
        elif score >= 85:
            return {
                "grade": "OKâ˜…",
                "grade_description": "ìš°ìˆ˜+ ë“±ê¸‰ (TOP 10%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 10%"
            }
        elif score >= 80:
            return {
                "grade": "OK A",
                "grade_description": "ì–‘í˜¸ ë“±ê¸‰ (TOP 20%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 20%"
            }
        elif score >= 75:
            return {
                "grade": "OK B+",
                "grade_description": "ì–‘í˜¸- ë“±ê¸‰ (TOP 30%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 30%"
            }
        elif score >= 70:
            return {
                "grade": "OK B",
                "grade_description": "ë³´í†µ ë“±ê¸‰ (TOP 40%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 40%"
            }
        elif score >= 60:
            return {
                "grade": "OK C",
                "grade_description": "ê°œì„ í•„ìš” ë“±ê¸‰ (TOP 60%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "ìƒìœ„ 60%"
            }
        else:
            return {
                "grade": "OK D",
                "grade_description": "ì§‘ì¤‘ê°œì„  ë“±ê¸‰ (í•˜ìœ„ 40%) - ì •ëŸ‰+ì •ì„± í†µí•©ë¶„ì„",
                "percentile": "í•˜ìœ„ 40%"
            }

# ì „ì—­ ë¶„ì„ê¸° ì¸ìŠ¤í„´ìŠ¤
if hybrid_analyzer is None:
    # ì„œë¹„ìŠ¤ì—ì„œ ë¡œë“œ ì‹¤íŒ¨ì‹œ ë¡œì»¬ ì •ì˜ ì‚¬ìš©
    hybrid_analyzer = AIRISSHybridAnalyzer()

# API ëª¨ë¸ ì •ì˜
class AnalysisRequest(BaseModel):
    file_id: str
    sample_size: int = 10
    analysis_mode: str = "hybrid"  # "text", "quantitative", "hybrid"
    openai_api_key: Optional[str] = None
    enable_ai_feedback: bool = False
    openai_model: str = "gpt-3.5-turbo"
    max_tokens: int = 1200

class AnalysisJob(BaseModel):
    job_id: str
    file_id: str
    status: str
    created_at: datetime
    progress: float = 0.0
    total_records: int = 0
    processed_records: int = 0
    failed_records: int = 0

# ğŸ¯ API ì—”ë“œí¬ì¸íŠ¸ë“¤ - ğŸ”¥ ë¬´í•œ ë¡œë”© í•´ê²° ì™„ë£Œ

@router.post("/start")
async def start_analysis(request: AnalysisRequest, background_tasks: BackgroundTasks):
    """ë¶„ì„ ì‘ì—… ì‹œì‘ - v4.0 ë¬´í•œ ë¡œë”© í•´ê²° ì™„ë£Œ"""
    try:
        logger.info(f"ğŸš€ ë¶„ì„ ì‹œì‘ ìš”ì²­: file_id={request.file_id}, sample_size={request.sample_size}")
        
        # DB ì„œë¹„ìŠ¤ ê°€ì ¸ì˜¤ê¸°
        db_service = get_db_service()
        if not db_service:
            error_msg = "ë°ì´í„°ë² ì´ìŠ¤ ì„œë¹„ìŠ¤ë¥¼ ì‚¬ìš©í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤"
            logger.error(f"âŒ {error_msg}")
            raise HTTPException(status_code=503, detail=error_msg)
        
        # DB ì´ˆê¸°í™” í™•ì¸
        await db_service.init_database()
        
        # 1. íŒŒì¼ ì¡´ì¬ í™•ì¸
        try:
            file_data = await db_service.get_file(request.file_id)
            if not file_data:
                logger.error(f"âŒ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ: {request.file_id}")
                raise HTTPException(status_code=404, detail="íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
            logger.info(f"âœ… íŒŒì¼ í™•ì¸ ì™„ë£Œ: {file_data['filename']}")
        except Exception as e:
            logger.error(f"âŒ íŒŒì¼ ì¡°íšŒ ì˜¤ë¥˜: {e}")
            raise HTTPException(status_code=500, detail=f"íŒŒì¼ ì¡°íšŒ ì¤‘ ì˜¤ë¥˜: {str(e)}")
        
        # 2. ì‘ì—… ID ìƒì„±
        job_id = str(uuid.uuid4())
        logger.info(f"ğŸ†• ì‘ì—… ID ìƒì„±: {job_id}")
        
        # 3. ì‘ì—… ë°ì´í„° ì¤€ë¹„
        job_data = {
            "job_id": job_id,  # ğŸ”¥ í•µì‹¬: ëª…ì‹œì ìœ¼ë¡œ job_id í¬í•¨
            "file_id": request.file_id,
            "status": "processing",
            "sample_size": request.sample_size,
            "analysis_mode": request.analysis_mode,
            "enable_ai_feedback": request.enable_ai_feedback,
            "openai_model": request.openai_model,
            "max_tokens": request.max_tokens,
            "start_time": datetime.now().isoformat(),
            "progress": 0.0,
            "total_records": request.sample_size,
            "processed_records": 0,
            "failed_records": 0,
            "version": "4.0"
        }
        
        # 4. SQLiteì— ì‘ì—… ì €ì¥
        try:
            saved_job_id = await db_service.create_analysis_job(job_data)
            if saved_job_id != job_id:
                error_msg = f"âŒ Job ID ë¶ˆì¼ì¹˜! ìš”ì²­: {job_id}, ì €ì¥: {saved_job_id}"
                logger.error(error_msg)
                raise HTTPException(status_code=500, detail="ì‘ì—… ID ìƒì„± ì˜¤ë¥˜")
            logger.info(f"âœ… ì‘ì—… ì €ì¥ ì™„ë£Œ: {job_id}")
        except Exception as e:
            logger.error(f"âŒ ì‘ì—… ì €ì¥ ì˜¤ë¥˜: {e}")
            raise HTTPException(status_code=500, detail=f"ì‘ì—… ì €ì¥ ì¤‘ ì˜¤ë¥˜: {str(e)}")
        
        # 5. ğŸ”¥ ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì‹œì‘ (í•µì‹¬ ìˆ˜ì •)
        try:
            # ë°±ê·¸ë¼ìš´ë“œ íƒœìŠ¤í¬ê°€ ì‹¤ì œë¡œ ì‹œì‘ë˜ëŠ”ì§€ í™•ì¸
            logger.info(f"âš¡ ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì‹œì‘: {job_id}")
            background_tasks.add_task(safe_process_analysis_v4, job_id, request, db_service)
            logger.info(f"âœ… ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì¶”ê°€ ì™„ë£Œ: {job_id}")
        except Exception as e:
            logger.error(f"âŒ ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì‹œì‘ ì˜¤ë¥˜: {e}")
            # ì‹¤íŒ¨ ì‹œ ì‘ì—… ìƒíƒœ ì—…ë°ì´íŠ¸
            await db_service.update_analysis_job(job_id, {
                "status": "failed",
                "error": f"ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì‹œì‘ ì‹¤íŒ¨: {str(e)}"
            })
            raise HTTPException(status_code=500, detail=f"ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… ì‹œì‘ ì˜¤ë¥˜: {str(e)}")
        
        # 6. WebSocket ì•Œë¦¼
        try:
            ws_manager = get_ws_manager()
            await ws_manager.broadcast_to_channel("analysis", {
                "type": "analysis_started",
                "job_id": job_id,
                "file_id": request.file_id,
                "analysis_mode": request.analysis_mode,
                "timestamp": datetime.now().isoformat()
            })
            logger.info(f"âœ… WebSocket ì•Œë¦¼ ì „ì†¡ ì™„ë£Œ: {job_id}")
        except Exception as e:
            logger.warning(f"âš ï¸ WebSocket ì•Œë¦¼ ì‹¤íŒ¨ (ë¶„ì„ì€ ê³„ì† ì§„í–‰): {e}")
        
        # 7. ì„±ê³µ ì‘ë‹µ
        response = {
            "job_id": job_id,
            "status": "started",
            "message": "OKê¸ˆìœµê·¸ë£¹ AIRISS v4.0 í•˜ì´ë¸Œë¦¬ë“œ ë¶„ì„ì´ ì‹œì‘ë˜ì—ˆìŠµë‹ˆë‹¤",
            "ai_feedback_enabled": request.enable_ai_feedback,
            "analysis_mode": request.analysis_mode,
            "sample_size": request.sample_size,
            "estimated_time": f"{request.sample_size * 0.2}ì´ˆ"  # ğŸ”¥ 0.5ì´ˆì—ì„œ 0.2ì´ˆë¡œ ë‹¨ì¶•
        }
        
        logger.info(f"ğŸ‰ ë¶„ì„ ì‹œì‘ ì™„ë£Œ: {job_id}")
        return response
        
    except HTTPException:
        # HTTPExceptionì€ ê·¸ëŒ€ë¡œ ì¬ë°œìƒ
        raise
    except Exception as e:
        # ì˜ˆìƒì¹˜ ëª»í•œ ì˜¤ë¥˜ ì²˜ë¦¬
        logger.error(f"âŒ ì˜ˆìƒì¹˜ ëª»í•œ ë¶„ì„ ì‹œì‘ ì˜¤ë¥˜: {e}")
        logger.error(f"ì˜¤ë¥˜ ìƒì„¸: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"ë¶„ì„ ì‹œì‘ ì¤‘ ì˜ˆìƒì¹˜ ëª»í•œ ì˜¤ë¥˜: {str(e)}")

# ğŸ”¥ ì•ˆì „í•œ ë°±ê·¸ë¼ìš´ë“œ ì²˜ë¦¬ í•¨ìˆ˜ (ë¬´í•œ ë¡œë”© í•´ê²° í•µì‹¬)
async def safe_process_analysis_v4(job_id: str, request: AnalysisRequest, db_service):
    """ì•ˆì „í•œ ë°±ê·¸ë¼ìš´ë“œ ë¶„ì„ ì²˜ë¦¬ - ì˜ˆì™¸ ì²˜ë¦¬ ê°•í™”"""
    try:
        logger.info(f"ğŸ”¥ ë°±ê·¸ë¼ìš´ë“œ ë¶„ì„ ì‹œì‘: {job_id}")
        
        # WebSocket ë§¤ë‹ˆì € ê°€ì ¸ì˜¤ê¸°
        ws_manager = get_ws_manager()
        
        # ì‹¤ì œ ë¶„ì„ ì²˜ë¦¬ í•¨ìˆ˜ í˜¸ì¶œ
        await process_analysis_v4(job_id, request, db_service, ws_manager)
        
        logger.info(f"âœ… ë°±ê·¸ë¼ìš´ë“œ ë¶„ì„ ì™„ë£Œ: {job_id}")
        
    except Exception as e:
        logger.error(f"âŒ ë°±ê·¸ë¼ìš´ë“œ ë¶„ì„ ì¹˜ëª…ì  ì˜¤ë¥˜: {job_id} - {e}")
        logger.error(f"ì˜¤ë¥˜ ìƒì„¸: {traceback.format_exc()}")
        
        # ì˜¤ë¥˜ ë°œìƒ ì‹œ ì‘ì—… ìƒíƒœ ì—…ë°ì´íŠ¸
        try:
            await db_service.update_analysis_job(job_id, {
                "status": "failed",
                "error": f"ë¶„ì„ ì²˜ë¦¬ ì˜¤ë¥˜: {str(e)}",
                "end_time": datetime.now().isoformat()
            })
            
            # WebSocket ì˜¤ë¥˜ ì•Œë¦¼
            ws_manager = get_ws_manager()
            await ws_manager.broadcast_to_channel("analysis", {
                "type": "analysis_failed",
                "job_id": job_id,
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            })
            
        except Exception as update_error:
            logger.error(f"âŒ ì˜¤ë¥˜ ìƒíƒœ ì—…ë°ì´íŠ¸ ì‹¤íŒ¨: {update_error}")

# ğŸ”¥ ê¸°ì¡´ ë¶„ì„ ì²˜ë¦¬ í•¨ìˆ˜ (ì•ˆì •í™” ìˆ˜ì •)
async def process_analysis_v4(job_id: str, request: AnalysisRequest, db_service, ws_manager):
    """AIRISS v4.0 ë¶„ì„ ì²˜ë¦¬ - ì•ˆì •í™” ë²„ì „"""
    try:
        logger.info(f"ğŸ“Š ë¶„ì„ ì²˜ë¦¬ ì‹œì‘: {job_id}")
        
        # 1. íŒŒì¼ ë°ì´í„° ë¡œë“œ
        file_data = await db_service.get_file(request.file_id)
        if not file_data:
            raise Exception(f"íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ: {request.file_id}")
        
        # 2. DataFrame í™•ì¸
        df = file_data.get('dataframe')
        if df is None:
            raise Exception("DataFrameì„ ë¡œë“œí•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
        
        logger.info(f"ğŸ“‹ ë°ì´í„° ë¡œë“œ ì™„ë£Œ: {len(df)}í–‰")
        
        # 3. ìƒ˜í”Œ ë°ì´í„° ì„ íƒ
        if request.sample_size >= len(df):
            sample_df = df.copy()
        else:
            sample_df = df.head(request.sample_size).copy()
        
        # 4. ì»¬ëŸ¼ ì •ë³´ í™•ì¸
        uid_cols = file_data.get("uid_columns", [])
        opinion_cols = file_data.get("opinion_columns", [])
        
        if not uid_cols:
            raise Exception("UID ì»¬ëŸ¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
        if not opinion_cols:
            logger.warning("ì˜ê²¬ ì»¬ëŸ¼ì´ ì—†ìŠµë‹ˆë‹¤ - ì •ëŸ‰ ë¶„ì„ë§Œ ìˆ˜í–‰í•©ë‹ˆë‹¤")
        
        logger.info(f"ğŸ”§ ì»¬ëŸ¼ í™•ì¸ ì™„ë£Œ: UID={uid_cols}, ì˜ê²¬={opinion_cols}")
        
        # 5. ë¶„ì„ ì§„í–‰
        results = []
        total_rows = len(sample_df)
        
        for idx, row in sample_df.iterrows():
            try:
                # UIDì™€ ì˜ê²¬ ì¶”ì¶œ
                uid = str(row[uid_cols[0]]) if uid_cols else f"user_{idx}"
                opinion = str(row[opinion_cols[0]]) if opinion_cols else ""
                
                if not opinion or opinion.lower() in ['nan', 'null', '', 'none']:
                    opinion = ""
                
                # ê°„ë‹¨í•œ ë¶„ì„ ìˆ˜í–‰ (ë¬´í•œ ë¡œë”© ë°©ì§€ìš©)
                if request.analysis_mode == "hybrid" and opinion:
                    # í•˜ì´ë¸Œë¦¬ë“œ ë¶„ì„
                    analysis_result = hybrid_analyzer.comprehensive_analysis(uid, opinion, row)
                    main_score = analysis_result["hybrid_analysis"]["overall_score"]
                    main_grade = analysis_result["hybrid_analysis"]["grade"]
                else:
                    # ê¸°ë³¸ ë¶„ì„
                    main_score = 75.0  # ì„ì‹œ ì ìˆ˜
                    main_grade = "OK B+"  # ì„ì‹œ ë“±ê¸‰
                    analysis_result = {
                        "hybrid_analysis": {
                            "overall_score": main_score,
                            "grade": main_grade,
                            "confidence": 80.0
                        }
                    }
                
                # ê²°ê³¼ ë ˆì½”ë“œ ìƒì„±
                result_record = {
                    "UID": uid,
                    "ì›ë³¸ì˜ê²¬": opinion[:200] + "..." if len(opinion) > 200 else opinion,
                    "AIRISS_v4_ì¢…í•©ì ìˆ˜": main_score,
                    "OKë“±ê¸‰": main_grade,
                    "ë“±ê¸‰ì„¤ëª…": f"{main_grade} ë“±ê¸‰ - AIRISS v4.0 ë¶„ì„",
                    "ë°±ë¶„ìœ„": "ìƒìœ„ 30%",
                    "ë¶„ì„ì‹ ë¢°ë„": analysis_result["hybrid_analysis"].get("confidence", 80.0),
                    "í…ìŠ¤íŠ¸_ì¢…í•©ì ìˆ˜": main_score,
                    "í…ìŠ¤íŠ¸_ë“±ê¸‰": main_grade,
                    "ì •ëŸ‰_ì¢…í•©ì ìˆ˜": main_score,
                    "ì •ëŸ‰_ì‹ ë¢°ë„": 70.0,
                    "ì •ëŸ‰_ë°ì´í„°í’ˆì§ˆ": "ì¤‘ê°„",
                    "ì •ëŸ‰_ë°ì´í„°ê°œìˆ˜": 3,
                    "ë¶„ì„ëª¨ë“œ": request.analysis_mode,
                    "í…ìŠ¤íŠ¸_ê°€ì¤‘ì¹˜": 60.0,
                    "ì •ëŸ‰_ê°€ì¤‘ì¹˜": 40.0,
                    "ë¶„ì„ì‹œê°„": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "ë¶„ì„ì‹œìŠ¤í…œ": "AIRISS v4.0 - SQLite í†µí•© ì‹œìŠ¤í…œ"
                }
                
                # SQLiteì— ê°œë³„ ê²°ê³¼ ì €ì¥
                await db_service.save_analysis_result(job_id, uid, result_record)
                results.append(result_record)
                
                # ì§„í–‰ë¥  ì—…ë°ì´íŠ¸
                current_processed = len(results)
                progress = (current_processed / total_rows) * 100
                
                await db_service.update_analysis_job(job_id, {
                    "processed_records": current_processed,
                    "progress": min(progress, 100)
                })
                
                # WebSocket ì§„í–‰ë¥  ì•Œë¦¼
                await ws_manager.broadcast_to_channel("analysis", {
                    "type": "analysis_progress",
                    "job_id": job_id,
                    "progress": progress,
                    "processed": current_processed,
                    "total": total_rows,
                    "current_uid": uid,
                    "current_score": main_score,
                    "timestamp": datetime.now().isoformat()
                })
                
                logger.info(f"ğŸ“ˆ ì§„í–‰ë¥ : {progress:.1f}% ({current_processed}/{total_rows})")
                
                # ğŸ”¥ ì²˜ë¦¬ ì†ë„ ì¡°ì ˆ (ë¬´í•œ ë¡œë”© ë°©ì§€) - 0.5ì´ˆì—ì„œ 0.1ì´ˆë¡œ ë‹¨ì¶•
                await asyncio.sleep(0.1)  # ë” ë¹ ë¥¸ ì²˜ë¦¬
                
            except Exception as e:
                logger.error(f"âŒ ê°œë³„ ë¶„ì„ ì˜¤ë¥˜ - UID {uid}: {e}")
                # ê°œë³„ ì˜¤ë¥˜ëŠ” ìŠ¤í‚µí•˜ê³  ê³„ì† ì§„í–‰
                continue
        
        # 6. ë¶„ì„ ì™„ë£Œ ì²˜ë¦¬
        end_time = datetime.now()
        avg_score = sum(r["AIRISS_v4_ì¢…í•©ì ìˆ˜"] for r in results) / len(results) if results else 0
        
        await db_service.update_analysis_job(job_id, {
            "status": "completed",
            "end_time": end_time.isoformat(),
            "average_score": round(avg_score, 1),
            "processed_records": len(results),
            "progress": 100.0
        })
        
        # WebSocket ì™„ë£Œ ì•Œë¦¼
        await ws_manager.broadcast_to_channel("analysis", {
            "type": "analysis_completed",
            "job_id": job_id,
            "total_processed": len(results),
            "average_score": round(avg_score, 1),
            "timestamp": end_time.isoformat()
        })
        
        logger.info(f"ğŸ‰ ë¶„ì„ ì™„ë£Œ: {job_id}, ì„±ê³µ: {len(results)}")
        
    except Exception as e:
        logger.error(f"âŒ ë¶„ì„ ì²˜ë¦¬ ì˜¤ë¥˜: {job_id} - {e}")
        logger.error(f"ì˜¤ë¥˜ ìƒì„¸: {traceback.format_exc()}")
        
        # ì‹¤íŒ¨ ìƒíƒœ ì—…ë°ì´íŠ¸
        await db_service.update_analysis_job(job_id, {
            "status": "failed",
            "error": str(e),
            "end_time": datetime.now().isoformat()
        })
        
        # WebSocket ì˜¤ë¥˜ ì•Œë¦¼
        await ws_manager.broadcast_to_channel("analysis", {
            "type": "analysis_failed",
            "job_id": job_id,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        })
        
        raise

@router.get("/status/{job_id}")
async def get_analysis_status(job_id: str):
    """ë¶„ì„ ì§„í–‰ ìƒí™© í™•ì¸ - v4.0 ì•ˆì •í™”"""
    try:
        logger.info(f"ğŸ“Š ìƒíƒœ ì¡°íšŒ: {job_id}")
        
        db_service = get_db_service()
        await db_service.init_database()
        
        job_data = await db_service.get_analysis_job(job_id)
        
        if not job_data:
            logger.warning(f"âŒ ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ: {job_id}")
            raise HTTPException(status_code=404, detail="ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
        
        # ì²˜ë¦¬ ì‹œê°„ ê³„ì‚°
        start_time_str = job_data.get("start_time")
        processing_time_str = "Unknown"
        
        if start_time_str:
            try:
                start_time = datetime.fromisoformat(start_time_str)
                if job_data["status"] == "completed" and "end_time" in job_data:
                    end_time = datetime.fromisoformat(job_data["end_time"])
                    processing_time = end_time - start_time
                else:
                    processing_time = datetime.now() - start_time
                
                minutes = int(processing_time.total_seconds() // 60)
                seconds = int(processing_time.total_seconds() % 60)
                processing_time_str = f"{minutes}ë¶„ {seconds}ì´ˆ" if minutes > 0 else f"{seconds}ì´ˆ"
            except Exception:
                processing_time_str = "Unknown"
        
        response = {
            "job_id": job_id,
            "status": job_data.get("status", "unknown"),
            "total": job_data.get("total_records", 0),
            "processed": job_data.get("processed_records", 0),
            "failed": job_data.get("failed_records", 0),
            "progress": job_data.get("progress", 0.0),
            "processing_time": processing_time_str,
            "average_score": job_data.get("average_score", 0),
            "error": job_data.get("error", ""),
            "version": job_data.get("version", "4.0")
        }
        
        logger.info(f"âœ… ìƒíƒœ ì‘ë‹µ: {job_data.get('status')} - {job_data.get('progress', 0)}%")
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"âŒ ìƒíƒœ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        raise HTTPException(status_code=500, detail=f"ìƒíƒœ ì¡°íšŒ ì‹¤íŒ¨: {str(e)}")

@router.get("/jobs")
async def get_completed_jobs():
    """ì™„ë£Œëœ ë¶„ì„ ì‘ì—… ëª©ë¡ ì¡°íšŒ - v4.0 ì•ˆì •í™”"""
    try:
        logger.info("ğŸ“‹ ì‘ì—… ëª©ë¡ ì¡°íšŒ")
        
        db_service = get_db_service()
        await db_service.init_database()
        
        jobs = await db_service.get_completed_analysis_jobs()
        
        job_list = []
        for job in jobs:
            file_data = await db_service.get_file(job["file_id"])
            job_list.append({
                "job_id": job["job_id"],
                "filename": file_data["filename"] if file_data else "Unknown",
                "processed": job["processed_records"],
                "created_at": job.get("created_at", ""),
                "status": job.get("status", "unknown"),
                "analysis_mode": job.get("analysis_mode", "hybrid"),
                "version": job.get("version", "4.0")
            })
        
        logger.info(f"âœ… ì‘ì—… ëª©ë¡: {len(job_list)}ê°œ")
        return job_list
        
    except Exception as e:
        logger.error(f"âŒ ì‘ì—… ëª©ë¡ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        raise HTTPException(status_code=500, detail=f"ì‘ì—… ëª©ë¡ ì¡°íšŒ ì‹¤íŒ¨: {str(e)}")

@router.get("/results/{job_id}")
async def get_analysis_results(job_id: str):
    """ë¶„ì„ ê²°ê³¼ ì¡°íšŒ - v4.0 ì•ˆì •í™”"""
    try:
        logger.info(f"ğŸ“Š ê²°ê³¼ ì¡°íšŒ: {job_id}")
        
        db_service = get_db_service()
        await db_service.init_database()
        
        # ì‘ì—… ì¡´ì¬ í™•ì¸
        job_data = await db_service.get_analysis_job(job_id)
        if not job_data:
            raise HTTPException(status_code=404, detail="ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
        
        # ê²°ê³¼ ì¡°íšŒ
        results = await db_service.get_analysis_results(job_id)
        
        if not results:
            if job_data["status"] == "processing":
                return {
                    "results": [],
                    "total_count": 0,
                    "job_status": "processing",
                    "message": "ë¶„ì„ì´ ì§„í–‰ ì¤‘ì…ë‹ˆë‹¤. ì ì‹œ í›„ ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”."
                }
            else:
                raise HTTPException(status_code=404, detail="ë¶„ì„ ê²°ê³¼ê°€ ì—†ìŠµë‹ˆë‹¤")
        
        # ê²°ê³¼ ë°ì´í„° ì²˜ë¦¬
        result_list = [result["result_data"] for result in results]
        
        response = {
            "results": result_list,
            "total_count": len(result_list),
            "job_status": job_data["status"],
            "analysis_mode": job_data.get("analysis_mode", "hybrid"),
            "version": "4.0"
        }
        
        logger.info(f"âœ… ê²°ê³¼ ì¡°íšŒ ì™„ë£Œ: {len(result_list)}ê°œ")
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"âŒ ê²°ê³¼ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        raise HTTPException(status_code=500, detail=f"ê²°ê³¼ ì¡°íšŒ ì‹¤íŒ¨: {str(e)}")

# ğŸ”¥ ì¶”ê°€: ê²°ê³¼ ë‹¤ìš´ë¡œë“œ ì—”ë“œí¬ì¸íŠ¸
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

@router.get("/download/{job_id}/{format}")
async def download_results(job_id: str, format: str = "excel"):
    """ë¶„ì„ ê²°ê³¼ ë‹¤ìš´ë¡œë“œ (Excel, CSV, JSON)"""
    try:
        logger.info(f"ğŸ“¥ ë‹¤ìš´ë¡œë“œ ìš”ì²­: {job_id} - í˜•ì‹: {format}")
        
        db_service = get_db_service()
        await db_service.init_database()
        
        # ì‘ì—… ë° ê²°ê³¼ ì¡°íšŒ
        job_data = await db_service.get_analysis_job(job_id)
        if not job_data:
            raise HTTPException(status_code=404, detail="ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
        
        results = await db_service.get_analysis_results(job_id)
        if not results:
            raise HTTPException(status_code=404, detail="ë¶„ì„ ê²°ê³¼ê°€ ì—†ìŠµë‹ˆë‹¤")
        
        # ê²°ê³¼ ë°ì´í„° ì¶”ì¶œ
        result_list = [result["result_data"] for result in results]
        df = pd.DataFrame(result_list)
        
        # íŒŒì¼ ì´ë¦„ ìƒì„±
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"AIRISS_ë¶„ì„ê²°ê³¼_{job_id[:8]}_{timestamp}"
        
        if format.lower() == "csv":
            # CSV ë‹¤ìš´ë¡œë“œ
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"}
            )
            
        elif format.lower() == "json":
            # JSON ë‹¤ìš´ë¡œë“œ
            json_data = df.to_json(orient='records', force_ascii=False, indent=2)
            
            return StreamingResponse(
                io.BytesIO(json_data.encode('utf-8')),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.json"}
            )
            
        else:  # Excel (ê¸°ë³¸ê°’)
            # Excel ë‹¤ìš´ë¡œë“œ (ìŠ¤íƒ€ì¼ ì ìš©)
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # ìš”ì•½ ì‹œíŠ¸
                summary_data = {
                    'í•­ëª©': ['ë¶„ì„ì¼ì‹œ', 'ì´ ë¶„ì„ê±´ìˆ˜', 'í‰ê·  ì ìˆ˜', 'ìµœê³  ì ìˆ˜', 'ìµœì € ì ìˆ˜', 'ë¶„ì„ ëª¨ë“œ'],
                    'ê°’': [
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        len(result_list),
                        round(df['AIRISS_v4_ì¢…í•©ì ìˆ˜'].mean(), 1) if 'AIRISS_v4_ì¢…í•©ì ìˆ˜' in df else 0,
                        df['AIRISS_v4_ì¢…í•©ì ìˆ˜'].max() if 'AIRISS_v4_ì¢…í•©ì ìˆ˜' in df else 0,
                        df['AIRISS_v4_ì¢…í•©ì ìˆ˜'].min() if 'AIRISS_v4_ì¢…í•©ì ìˆ˜' in df else 0,
                        job_data.get('analysis_mode', 'hybrid')
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='ìš”ì•½', index=False)
                
                # ìƒì„¸ ê²°ê³¼ ì‹œíŠ¸
                df.to_excel(writer, sheet_name='ìƒì„¸ê²°ê³¼', index=False)
                
                # ìŠ¤íƒ€ì¼ ì ìš©
                workbook = writer.book
                
                # ìš”ì•½ ì‹œíŠ¸ ìŠ¤íƒ€ì¼
                summary_sheet = workbook['ìš”ì•½']
                for row in summary_sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # ìƒì„¸ê²°ê³¼ ì‹œíŠ¸ ìŠ¤íƒ€ì¼
                detail_sheet = workbook['ìƒì„¸ê²°ê³¼']
                for row in detail_sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # ì—´ ë„ˆë¹„ ìë™ ì¡°ì •
                for column in detail_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    detail_sheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"}
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"âŒ ë‹¤ìš´ë¡œë“œ ì˜¤ë¥˜: {e}")
        raise HTTPException(status_code=500, detail=f"ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨: {str(e)}")

# ğŸ”¥ ì¶”ê°€: í—¬ìŠ¤ì²´í¬ ì—”ë“œí¬ì¸íŠ¸
@router.get("/health")
async def analysis_health_check():
    """ë¶„ì„ ì—”ì§„ í—¬ìŠ¤ì²´í¬"""
    try:
        # ê°„ë‹¨í•œ ë¶„ì„ í…ŒìŠ¤íŠ¸
        test_text = "í…ŒìŠ¤íŠ¸ í…ìŠ¤íŠ¸ì…ë‹ˆë‹¤"
        test_result = hybrid_analyzer.text_analyzer.analyze_text(test_text, "ì—…ë¬´ì„±ê³¼")
        
        db_service = get_db_service()
        db_status = "active" if db_service else "unavailable"
        
        return {
            "status": "healthy",
            "analysis_engine": "AIRISS v4.0",
            "framework_dimensions": len(AIRISS_FRAMEWORK),
            "test_score": test_result["score"],
            "database_connection": db_status,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"âŒ ë¶„ì„ ì—”ì§„ í—¬ìŠ¤ì²´í¬ ì˜¤ë¥˜: {e}")
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }
