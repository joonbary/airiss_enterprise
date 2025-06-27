# app/core/excel_reporter.py
"""Excel 리포트 생성 엔진"""

import pandas as pd
from datetime import datetime
import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.core.analyzers.framework import AIRISS_FRAMEWORK

class ExcelReporter:
    def __init__(self):
        self.output_dir = "results"
        os.makedirs(self.output_dir, exist_ok=True)
    
    async def create_report(
        self,
        job_id: str,
        results: List[Dict[str, Any]],
        analysis_mode: str = "hybrid",
        enable_ai: bool = False
    ) -> str:
        """Excel 보고서 생성"""
        
        # 데이터프레임 생성
        df_results = pd.DataFrame(results)
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ai_suffix = "_AI피드백포함" if enable_ai else ""
        filename = f"AIRISS_v4.0_{analysis_mode}{ai_suffix}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Excel Writer 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 메인 결과 시트
            df_results.to_excel(writer, sheet_name='분석결과', index=False)
            
            # 2. 통계 요약 시트
            summary_df = self._create_summary(df_results)
            summary_df.to_excel(writer, sheet_name='통계요약', index=False)
            
            # 3. 8대 영역별 분석 시트
            dimension_df = self._create_dimension_analysis(df_results)
            dimension_df.to_excel(writer, sheet_name='영역별분석', index=False)
            
            # 워크북 스타일링
            workbook = writer.book
            self._style_workbook(workbook)
        
        return filepath
    
    def _create_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """통계 요약 생성"""
        summary_data = []
        
        # 기본 통계
        summary_data.append({
            "항목": "전체 분석 건수",
            "값": len(df),
            "설명": "총 분석된 직원 수"
        })
        
        if "hybrid_score" in df.columns:
            summary_data.append({
                "항목": "평균 하이브리드 점수",
                "값": round(df["hybrid_score"].mean(), 1),
                "설명": "전체 직원 평균 통합 점수"
            })
        
        # OK등급 분포
        if "ok_grade" in df.columns:
            grade_counts = df["ok_grade"].value_counts()
            for grade, count in grade_counts.items():
                percentage = (count / len(df)) * 100
                summary_data.append({
                    "항목": f"{grade} 등급",
                    "값": f"{count}명 ({percentage:.1f}%)",
                    "설명": f"{grade} 등급 직원 수"
                })
        
        return pd.DataFrame(summary_data)
    
    def _create_dimension_analysis(self, df: pd.DataFrame) -> pd.DataFrame:
        """8대 영역별 분석"""
        dimension_data = []
        
        for dimension, info in AIRISS_FRAMEWORK.items():
            dim_col = f"{dimension}_score"
            if dim_col in df.columns:
                scores = df[dim_col]
                dimension_data.append({
                    "영역": dimension,
                    "아이콘": info["icon"],
                    "가중치": f"{info['weight']*100:.0f}%",
                    "설명": info["description"],
                    "평균점수": round(scores.mean(), 1),
                    "최고점수": round(scores.max(), 1),
                    "최저점수": round(scores.min(), 1),
                    "표준편차": round(scores.std(), 1)
                })
        
        return pd.DataFrame(dimension_data)
    
    def _style_workbook(self, workbook):
        """Excel 스타일 적용"""
        # OK금융그룹 브랜드 컬러
        ok_orange = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
        ok_grey = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # 헤더 스타일
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 각 시트 스타일링
        for sheet in workbook.worksheets:
            # 헤더 행 스타일
            for cell in sheet[1]:
                cell.fill = ok_orange
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 컬럼 너비 자동 조정
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width